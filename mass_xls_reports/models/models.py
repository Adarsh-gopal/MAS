# -*- coding: utf-8 -*-

import io
import locale
import base64
import textwrap
from copy import copy
from openpyxl import Workbook
from odoo import models, fields, api,_
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from odoo.exceptions import UserError, ValidationError,Warning
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from PIL import Image as PILImage
import base64
# import tempfile
# import os

from io import StringIO
class SaleOrder(models.Model):
    _inherit = 'sale.order'

    sale_quotation_xlsx_report = fields.Binary()
    file_name = fields.Char()

    with_pictures = fields.Boolean()


    def generate_report(self):
        #Initialize Constraints
        disc_adj = 3
        for line in self.order_line:
            if line.discount:
                disc_adj = 0
                break

        
        #Create Workbook and Worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Quotation"

        
        #Image placement settings
        c2e = cm_to_EMU
        p2e = pixels_to_EMU
        cellh = lambda x: c2e((x * 49.77)/99)
        cellw = lambda x: c2e((x * (18.65-1.71))/10)
        coloffset = cellw(0.1)
        rowoffset = cellh(0.1)
        color1 = PatternFill(start_color='C4BC96',
                   end_color='C4BC96',
                   fill_type='solid')
        color2 = PatternFill(start_color='95B3D7',
                   end_color='95B3D7',
                   fill_type='solid')
        color3 = PatternFill(start_color='92D050',
                   end_color='92D050',
                   fill_type='solid')
        color4 = PatternFill(start_color='C4D69B',end_color='C4D69B',
                   fill_type='solid')
        color5 = PatternFill(start_color='FF0000',end_color='FF0000',
                   fill_type='solid')



        #Border
        thin = Side(style="thin", color="000000")
        
        #Header Section
        #Logo



        if self.material_group == 'door':
            decoded_img = base64.b64decode(self.env.company.logo)
            img = Image(io.BytesIO(decoded_img))
            # h, w = img.height, img.width
            h, w = 94, 132
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            marker = AnchorMarker(col=0, colOff=coloffset, row=0, rowOff=rowoffset)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(img)
            logo = ws['A1']
            logo.border = Border(bottom=thin,left=thin,top=thin)
            ws.merge_cells(start_row=1, start_column=1, end_row=5, end_column=2)
            
            #Address
            address = ws.cell(row=1, column=3, value="#36, Mysore-Hunsur road,\nHootagalli, Belavadi Post,\nMysore -570018 India\nTel : +91 821 2404242\nFax : +91 821 2404343\ninfo@masfurn.com")
            address.font = Font(size=10,name='Calibri')
            address.alignment = Alignment(horizontal='center',vertical='center')
            address.border = Border(bottom=thin,top=thin)
            ws.merge_cells(start_row=1, start_column=3, end_row=5, end_column=5)
            
            #Title
            title = ws.cell(row=1, column=6, value="QUOTATION")
            title.font = Font(size=20,bold=True,name='Calibri')
            title.alignment = copy(address.alignment)
            title.border = copy(address.border)
            ws.merge_cells(start_row=1, start_column=6, end_row=5, end_column=20 - disc_adj)

            #HeadInfoRight
            #Quotation Number
            quotation_no = ws.cell(row=1,column=21 - disc_adj,value="QUOTATION NO: "+self.name)
            quotation_no.font = Font(size=10,name='Calibri')
            quotation_no.alignment = copy(address.alignment)
            quotation_no.border = Border(top=thin,right=thin)
            ws.merge_cells(start_row=1, start_column=21 - disc_adj, end_row=2, end_column=23 - disc_adj)
            #Date
            date = ws.cell(row=3,column=21 - disc_adj,value="Date: "+str(self.date_order))
            date.font = Font(size=10,name='Calibri')
            date.alignment = copy(address.alignment)
            date.border = Border(right=thin)
            ws.merge_cells(start_row=3, start_column=21 - disc_adj, end_row=3, end_column=23 - disc_adj)
            #Validity
            if self.validity_date:
                valid_days = str((self.validity_date - self.date_order.date()).days)
                validity = ws.cell(row=4, column=21 - disc_adj,value="Quote valid for "+valid_days+" days")
                validity.font = Font(size=10,name='Calibri',color=colors.RED)
                validity.alignment = copy(address.alignment)
                validity.border = Border(bottom=thin,right=thin)
            else:
                validity = ws.cell(row=4, column=21 - disc_adj)
                validity.border = Border(bottom=thin,right=thin)
            ws.merge_cells(start_row=4, start_column=21 - disc_adj, end_row=5, end_column=23 - disc_adj)


            #Analytic Account
            analytic = ws.cell(row=6, column=1, value=self.analytic_account_id.name)
            analytic.font = Font(size=15,name='Calibri',bold=True)
            analytic.alignment = copy(address.alignment)
            analytic.fill = color4
            analytic.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=6, start_column=1, end_row=7, end_column=23 - disc_adj)
            #column headings for material type door
            
            #serial number
            serial_no = ws.cell(row=8, column=1, value="Serial\nNumber")
            serial_no.font = Font(size=10,name='Calibri',bold=True)
            serial_no.alignment = copy(address.alignment)
            serial_no.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=1, end_row=9, end_column=1)

            #description
            description = ws.cell(row=8, column=2, value="Description")
            description.font = copy(serial_no.font)
            description.alignment = copy(address.alignment)
            description.border = copy(serial_no.border)
            ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=8)

            #Location
            location = ws.cell(row=9, column=2, value="Location")
            location.font = copy(serial_no.font)
            location.alignment = copy(address.alignment)
            location.border = copy(serial_no.border)
            #Face
            face = ws.cell(row=9, column=3, value="Face")
            face.font = copy(serial_no.font)
            face.alignment = copy(address.alignment)
            face.border = copy(serial_no.border)
            ws.merge_cells(start_row=9, start_column=3, end_row=9, end_column=4)
            #Back
            back = ws.cell(row=9, column=5, value="Back")
            back.font = copy(serial_no.font)
            back.alignment = copy(address.alignment)
            back.border = copy(serial_no.border)
            ws.merge_cells(start_row=9, start_column=5, end_row=9, end_column=6)
            #Edge Banding
            edge_banding = ws.cell(row=9, column=7, value="Edge Banding")
            edge_banding.font = copy(serial_no.font)
            edge_banding.alignment = copy(address.alignment)
            edge_banding.border = copy(serial_no.border)
            ws.merge_cells(start_row=9, start_column=7, end_row=9, end_column=8)



            #Door Size
            door_size = ws.cell(row=8, column=9, value="Door Size")
            door_size.font = copy(serial_no.font)
            door_size.alignment = copy(address.alignment)
            door_size.border = copy(serial_no.border)
            ws.merge_cells(start_row=8, start_column=9, end_row=8, end_column=13)

            #Length
            length = ws.cell(row=9, column=9, value="L(mm)")
            length.font = copy(serial_no.font)
            length.alignment = copy(address.alignment)
            length.border = copy(serial_no.border)
            ws.column_dimensions['I'].width = ws.column_dimensions['I'].width*0.5
            #X
            x1 = ws.cell(row=9, column=10, value="x")
            x1.font = copy(serial_no.font)
            x1.alignment = copy(address.alignment)
            x1.border = copy(serial_no.border)
            ws.column_dimensions['J'].width = ws.column_dimensions['J'].width*0.25
            #Width
            width = ws.cell(row=9, column=11, value="W(mm)")
            width.font = copy(serial_no.font)
            width.alignment = copy(address.alignment)
            width.border = copy(serial_no.border)
            ws.column_dimensions['K'].width = ws.column_dimensions['K'].width*0.5
            #X
            x2 = ws.cell(row=9, column=12, value="x")
            x2.font = copy(serial_no.font)
            x2.alignment = copy(address.alignment)
            x2.border = copy(serial_no.border)
            ws.column_dimensions['L'].width = ws.column_dimensions['L'].width*0.25
            #Thickness
            thickness = ws.cell(row=9, column=13, value="T(mm)")
            thickness.font = copy(serial_no.font)
            thickness.alignment = copy(address.alignment)
            thickness.border = copy(serial_no.border)
            ws.column_dimensions['M'].width = ws.column_dimensions['M'].width*0.5

            #Basic Price
            basic_price = ws.cell(row=8, column=14, value="Basic\nPrice")
            basic_price.font = copy(serial_no.font)
            basic_price.alignment = copy(address.alignment)
            basic_price.border = copy(serial_no.border)
            ws.merge_cells(start_row=8, start_column=14, end_row=9, end_column=14)
            

            #Discount
            if disc_adj == 0:
                discount = ws.cell(row=8, column=15, value="Discount")
                discount.font = copy(serial_no.font)
                discount.alignment = copy(address.alignment)
                discount.border = copy(serial_no.border)
                ws.merge_cells(start_row=8, start_column=15, end_row=8, end_column=16)
                #Percent
                percent = ws.cell(row=9, column=15, value="( % )")
                percent.font = copy(serial_no.font)
                percent.alignment = copy(address.alignment)
                percent.border = copy(serial_no.border)
                #Amount
                amount = ws.cell(row=9, column=16, value="( Amt )")
                amount.font = copy(serial_no.font)
                amount.alignment = copy(address.alignment)
                amount.border = copy(serial_no.border)
                #Discounted Price
                discounted_price = ws.cell(row=8, column=17, value="Discounted\nPrice")
                discounted_price.font = copy(serial_no.font)
                discounted_price.alignment = copy(address.alignment)
                discounted_price.border = copy(serial_no.border)
                ws.merge_cells(start_row=8, start_column=17, end_row=9, end_column=17)
                ws.column_dimensions['Q'].width = ws.column_dimensions['Q'].width*0.8

            
            #QTY
            qty = ws.cell(row=8, column=18 - disc_adj, value="QTY")
            qty.font = copy(serial_no.font)
            qty.alignment = copy(address.alignment)
            qty.border = copy(serial_no.border)
            ws.merge_cells(start_row=8, start_column=18 - disc_adj, end_row=9, end_column=18 - disc_adj)
            ws.column_dimensions[chr(ord('R')-disc_adj)].width = ws.column_dimensions[chr(ord('R')-disc_adj)].width*0.5

            #Total Basic Price
            total_basic_price = ws.cell(row=8, column=19 - disc_adj, value="Total Basic\nPrice")
            total_basic_price.font = copy(serial_no.font)
            total_basic_price.alignment = copy(address.alignment)
            total_basic_price.border = copy(serial_no.border)
            ws.merge_cells(start_row=8, start_column=19 - disc_adj, end_row=9, end_column=19 - disc_adj)
            ws.column_dimensions[chr(ord('S')-disc_adj)].width = ws.column_dimensions[chr(ord('S')-disc_adj)].width*0.8

            #GST
            gst = ws.cell(row=8, column=20 - disc_adj, value="GST")
            gst.font = copy(serial_no.font)
            gst.alignment = copy(address.alignment)
            gst.border = copy(serial_no.border)
            ws.merge_cells(start_row=8, start_column=20 - disc_adj, end_row=8, end_column=21 - disc_adj)
            #Percent
            percent = ws.cell(row=9, column=20 - disc_adj, value="( % )")
            percent.font = copy(serial_no.font)
            percent.alignment = copy(address.alignment)
            percent.border = copy(serial_no.border)
            #Amount
            amount = ws.cell(row=9, column=21 - disc_adj, value="( Amt )")
            amount.font = copy(serial_no.font)
            amount.alignment = copy(address.alignment)
            amount.border = copy(serial_no.border)

            #Total Value Inclusive of Taxes
            total_value_with_tax = ws.cell(row=8, column=22 - disc_adj, value="Total Value\nInclusive of Taxes")
            total_value_with_tax.font = copy(serial_no.font)
            total_value_with_tax.alignment = copy(address.alignment)
            total_value_with_tax.border = copy(serial_no.border)
            ws.merge_cells(start_row=8, start_column=22 - disc_adj, end_row=9, end_column=23 - disc_adj)






            #Lines
            total_amount = 0
            current_row = 10
            sl_no = 1
            locale.setlocale(locale.LC_MONETARY, 'en_IN')
            def format(my_str,limit):
                wrapper = textwrap.TextWrapper(width=limit)
                string = wrapper.fill(text=my_str)
                return string
            
            for line in self.order_line:
                #Serial Number
                serial_no = ws.cell(row=current_row, column=1, value=sl_no)
                serial_no.font = Font(size=10,name='Calibri')
                serial_no.alignment = copy(address.alignment)
                serial_no.border = Border(bottom=thin,top=thin,right=thin,left=thin)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+3, end_column=1)
                #Location
                location = ws.cell(row=current_row, column=2, value=line.product_id.name)
                location.font = Font(size=7,name='Calibri')
                location.alignment = copy(address.alignment)
                location.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row+3, end_column=2)
                #Face
                face = ws.cell(row=current_row, column=3, value=line.face_description)
                face.font = copy(location.font)
                face.alignment = copy(address.alignment)
                face.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row+3, end_column=4)
                #Back
                back = ws.cell(row=current_row, column=5, value=line.back_description)
                back.font = copy(location.font)
                back.alignment = copy(address.alignment)
                back.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row+3, end_column=6)
                #Edge Banding
                edge_banding = ws.cell(row=current_row, column=7, value=line.edge_banding)
                edge_banding.font = copy(location.font)
                edge_banding.alignment = copy(address.alignment)
                edge_banding.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row+3, end_column=8)
                #Length
                length = ws.cell(row=current_row, column=9, value=line.lengthx)
                length.font = copy(location.font)
                length.alignment = copy(address.alignment)
                length.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=9, end_row=current_row+3, end_column=9)
                #X
                x1 = ws.cell(row=current_row, column=10, value="x")
                x1.font = copy(location.font)
                x1.alignment = copy(address.alignment)
                x1.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=10, end_row=current_row+3, end_column=10)
                #Width
                width = ws.cell(row=current_row, column=11, value=line.width)
                width.font = copy(location.font)
                width.alignment = copy(address.alignment)
                width.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=11, end_row=current_row+3, end_column=11)
                #X
                x2 = ws.cell(row=current_row, column=12, value="x")
                x2.font = copy(location.font)
                x2.alignment = copy(address.alignment)
                x2.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=12, end_row=current_row+3, end_column=12)
                #Thickness
                thickness = ws.cell(row=current_row, column=13, value=line.thickness)
                thickness.font = copy(location.font)
                thickness.alignment = copy(address.alignment)
                thickness.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=13, end_row=current_row+3, end_column=13)
                #Basic Price
                basic_price = ws.cell(row=current_row, column=14, value=locale.currency(line.price_unit,grouping=True))
                basic_price.font = copy(location.font)
                basic_price.alignment = copy(address.alignment)
                basic_price.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=14, end_row=current_row+3, end_column=14)


                price = line.price_unit
                #Discount
                if disc_adj == 0:
                    discount = line.price_unit*line.discount/100
                    price = line.price_unit - discount
                    #Percent
                    percent = ws.cell(row=current_row, column=15, value=str(line.discount)+'%')
                    percent.font = copy(location.font)
                    percent.alignment = copy(address.alignment)
                    percent.border = copy(serial_no.border)
                    ws.merge_cells(start_row=current_row, start_column=15, end_row=current_row+3, end_column=15)
                    #Amount
                    amount = ws.cell(row=current_row, column=16, value=locale.currency(discount,grouping=True))
                    amount.font = copy(location.font)
                    amount.alignment = copy(address.alignment)
                    amount.border = copy(serial_no.border)
                    ws.merge_cells(start_row=current_row, start_column=16, end_row=current_row+3, end_column=16)
                    #Discounted Price
                    discounted_price = ws.cell(row=current_row, column=17, value=locale.currency(price,grouping=True))
                    discounted_price.font = copy(location.font)
                    discounted_price.alignment = copy(address.alignment)
                    discounted_price.border = copy(serial_no.border)
                    ws.merge_cells(start_row=current_row, start_column=17, end_row=current_row+3, end_column=17)


                #QTY
                qty = ws.cell(row=current_row, column=18 - disc_adj, value=line.product_uom_qty)
                qty.font = copy(location.font)
                qty.alignment = copy(address.alignment)
                qty.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=18 - disc_adj, end_row=current_row+3, end_column=18 - disc_adj)
                #Total Basic Price
                price = price * line.product_uom_qty
                total_basic_price = ws.cell(row=current_row, column=19 - disc_adj, value=locale.currency(price,grouping=True))
                total_basic_price.font = copy(location.font)
                total_basic_price.alignment = copy(address.alignment)
                total_basic_price.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=19 - disc_adj, end_row=current_row+3, end_column=19 - disc_adj)

                
                #Tax
                tax_prcnt = sum([sum([child.amount for child in tax.children_tax_ids]) for tax in line.tax_id])
                tax = price * tax_prcnt / 100
                price += tax
                #Percent
                percent = ws.cell(row=current_row, column=20 - disc_adj, value=str(tax_prcnt)+'%')
                percent.font = copy(location.font)
                percent.alignment = copy(address.alignment)
                percent.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=20 - disc_adj, end_row=current_row+3, end_column=20 - disc_adj)
                #Amount
                amount = ws.cell(row=current_row, column=21 - disc_adj, value=locale.currency(tax,grouping=True))
                amount.font = copy(location.font)
                amount.alignment = copy(address.alignment)
                amount.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=21 - disc_adj, end_row=current_row+3, end_column=21 - disc_adj)

                #Total Value Inclusive of Taxes
                total_value_with_tax = ws.cell(row=current_row, column=22 - disc_adj, value=locale.currency(price,grouping=True))
                total_value_with_tax.font = copy(location.font)
                total_value_with_tax.alignment = copy(address.alignment)
                total_value_with_tax.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=22 - disc_adj, end_row=current_row+3, end_column=23 - disc_adj)

                #update counters
                total_amount += price
                current_row += 4
                sl_no += 1


            #Grand Total
            grand_total = ws.cell(row=current_row, column=1, value="Grand Total")
            grand_total.font = Font(size=8,name='Calibri',bold=True)
            grand_total.alignment = Alignment(horizontal='right',vertical='center')
            grand_total.border = copy(serial_no.border)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=21 - disc_adj)
            #Grand Total Amount
            grand_total_amount = ws.cell(row=current_row, column=22 - disc_adj, value=locale.currency(total_amount,grouping=True))
            grand_total_amount.font = copy(grand_total.font)
            grand_total_amount.alignment = Alignment(horizontal='center',vertical='center')
            grand_total_amount.border = copy(serial_no.border)
            ws.merge_cells(start_row=current_row, start_column=22 - disc_adj, end_row=current_row, end_column=23 - disc_adj)
            #Footer
            footer = ws.cell(row=current_row+1, column=1, value="TAXES: Taxes as applicable at time of dispatch\nTRANSPORT:  Extra at actual\nFixing: rates do not include fixing charges\nPAYMENT: 50% advance balance 50% against delivery within 15 days from date of supply.")
            footer.font = Font(size=10,name='Calibri',bold=True)
            footer.alignment = Alignment(horizontal='left',vertical='center')
            footer.border = copy(serial_no.border)
            ws.merge_cells(start_row=current_row+1, start_column=1, end_row=current_row+4, end_column=23 - disc_adj)


        #door frame type
        if self.material_group == 'doorframe':
            # img = Image('/odoo14/prixgen_odoo14/mass_xls_reports/models/mas.jpg')
            decoded_img = base64.b64decode(self.env.company.logo)
            img = Image(io.BytesIO(decoded_img))
            # h, w = img.height, img.width
            h, w = 94, 132
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            marker = AnchorMarker(col=0, colOff=coloffset, row=0, rowOff=rowoffset)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(img)
            logo = ws['A1']
            logo.border = Border(bottom=thin,left=thin,top=thin)
            ws.merge_cells(start_row=1, start_column=1, end_row=5, end_column=2)
            
            #Address
            address = ws.cell(row=1, column=3, value="#36, Mysore-Hunsur road,\nHootagalli, Belavadi Post,\nMysore -570018 India\nTel : +91 821 2404242\nFax : +91 821 2404343\ninfo@masfurn.com")
            address.font = Font(size=10,name='Calibri')
            address.alignment = Alignment(horizontal='center',vertical='center')
            address.border = Border(bottom=thin,top=thin)
            ws.merge_cells(start_row=1, start_column=3, end_row=5, end_column=5)
            
            #Title
            title = ws.cell(row=1, column=6, value="QUOTATION")
            title.font = Font(size=20,bold=True,name='Calibri')
            title.alignment = copy(address.alignment)
            title.border = copy(address.border)
            ws.merge_cells(start_row=1, start_column=6, end_row=5, end_column=20 - disc_adj)

            #HeadInfoRight
            #Quotation Number
            quotation_no = ws.cell(row=1,column=21 - disc_adj,value="QUOTATION NO: "+self.name)
            quotation_no.font = Font(size=10,name='Calibri')
            quotation_no.alignment = copy(address.alignment)
            quotation_no.border = Border(top=thin,right=thin)
            ws.merge_cells(start_row=1, start_column=21 - disc_adj, end_row=2, end_column=23 - disc_adj)
            #Date
            date = ws.cell(row=3,column=21 - disc_adj,value="Date: "+str(self.date_order))
            date.font = Font(size=10,name='Calibri')
            date.alignment = copy(address.alignment)
            date.border = Border(right=thin)
            ws.merge_cells(start_row=3, start_column=21 - disc_adj, end_row=3, end_column=23 - disc_adj)
            #Validity
            if self.validity_date:
                valid_days = str((self.validity_date - self.date_order.date()).days)
                validity = ws.cell(row=4, column=21 - disc_adj,value="Quote valid for "+valid_days+" days")
                validity.font = Font(size=10,name='Calibri',color=colors.RED)
                validity.alignment = copy(address.alignment)
                validity.border = Border(bottom=thin,right=thin)
            else:
                validity = ws.cell(row=4, column=21 - disc_adj)
                validity.border = Border(bottom=thin,right=thin)
            ws.merge_cells(start_row=4, start_column=21 - disc_adj, end_row=5, end_column=23 - disc_adj)


            #Analytic Account
            analytic = ws.cell(row=6, column=1, value=self.analytic_account_id.name)
            analytic.font = Font(size=15,name='Calibri',bold=True)
            analytic.alignment = copy(address.alignment)
            analytic.fill = color4
            analytic.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=6, start_column=1, end_row=7, end_column=23 - disc_adj)


            #serial number
            serial_no = ws.cell(row=8, column=1, value="SR.\nNo.")
            serial_no.font = Font(size=10,name='Calibri',bold=True)
            serial_no.alignment = copy(address.alignment)
            serial_no.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=1, end_row=10, end_column=1)
            ws.column_dimensions['A'].width = ws.column_dimensions['A'].width*0.5

            #name
            name = ws.cell(row=8, column=2, value="Name")
            name.font = Font(size=10,name='Calibri',bold=True)
            name.alignment = copy(address.alignment)
            name.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=2, end_row=10, end_column=2)

            #Frame
            frame = ws.cell(row=8, column=3, value="Frame")
            frame.font = Font(size=10,name='Calibri',bold=True)
            frame.alignment = copy(address.alignment)
            frame.fill = color1
            frame.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=3, end_row=8, end_column=8)

            #under frame

            frame_desc = ws.cell(row=9, column=3, value="Frame Description")
            frame_desc.font = Font(size=10,name='Calibri',bold=True)
            frame_desc.alignment = copy(address.alignment)
            frame_desc.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=9, start_column=3, end_row=10, end_column=3)
            ws.column_dimensions['C'].width = ws.column_dimensions['C'].width * 2

            #frameoutersize
            frame_outersize = ws.cell(row=9, column=4, value="Frame Outer Size")
            frame_outersize.font = Font(size=10,name='Calibri',bold=True)
            frame_outersize.alignment = copy(address.alignment)
            # frame_outersize.set_bg_color('#C4BC96')
            frame_outersize.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=9, start_column=4, end_row=9, end_column=7)

            #L
            frame_outersize_length = ws.cell(row=10, column=4, value="L")
            frame_outersize_length.font = Font(size=10,name='Calibri',bold=True)
            frame_outersize_length.alignment = copy(address.alignment)
            frame_outersize_length.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['D'].width = ws.column_dimensions['D'].width*0.5
            # ws.merge_cells(start_row=9, start_column=4, end_row=9, end_column=7)

            #W
            frame_outersize_width = ws.cell(row=10, column=5, value="W")
            frame_outersize_width.font = Font(size=10,name='Calibri',bold=True)
            frame_outersize_width.alignment = copy(address.alignment)
            frame_outersize_width.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['E'].width = ws.column_dimensions['E'].width*0.5

            #D
            frame_outersize_diameter = ws.cell(row=10, column=6, value="D")
            frame_outersize_diameter.font = Font(size=10,name='Calibri',bold=True)
            frame_outersize_diameter.alignment = copy(address.alignment)
            frame_outersize_diameter.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['F'].width = ws.column_dimensions['F'].width*0.5

            #T
            frame_outersize_thickness = ws.cell(row=10, column=7, value="T")
            frame_outersize_thickness.font = Font(size=10,name='Calibri',bold=True)
            frame_outersize_thickness.alignment = copy(address.alignment)
            frame_outersize_thickness.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['G'].width = ws.column_dimensions['G'].width*0.5

            #Basic Rate/ frame
            frame_basic_rate = ws.cell(row=9, column=8, value="Basic Rate/\nFrame")
            frame_basic_rate.font = Font(size=10,name='Calibri',bold=True)
            frame_basic_rate.alignment = copy(address.alignment)
            frame_basic_rate.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=9, start_column=8, end_row=10, end_column=8)


            #Shutter
            shutter = ws.cell(row=8, column=9, value="Shutter")
            shutter.font = Font(size=10,name='Calibri',bold=True)
            shutter.alignment = copy(address.alignment)
            shutter.fill = color2
            shutter.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=9, end_row=8, end_column=13)

            #under frame
            door_desc = ws.cell(row=9, column=9, value="Door Description")
            door_desc.font = Font(size=10,name='Calibri',bold=True)
            door_desc.alignment = copy(address.alignment)
            door_desc.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=9, start_column=9, end_row=10, end_column=9)
            ws.column_dimensions['I'].width = ws.column_dimensions['I'].width * 2

            #frameoutersize
            door_final_size = ws.cell(row=9, column=10, value="Door Final Size")
            door_final_size.font = Font(size=10,name='Calibri',bold=True)
            door_final_size.alignment = copy(address.alignment)
            # door_final_size.set_bg_color('#C9BC96')
            door_final_size.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=9, start_column=10, end_row=9, end_column=12)

            #L
            shutter_length = ws.cell(row=10, column=10, value="L")
            shutter_length.font = Font(size=10,name='Calibri',bold=True)
            shutter_length.alignment = copy(address.alignment)
            shutter_length.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['J'].width = ws.column_dimensions['J'].width*0.5
            # ws.merge_cells(start_row=9, start_column=4, end_row=9, end_column=7)

            #W
            shutter_width = ws.cell(row=10, column=11, value="W")
            shutter_width.font = Font(size=10,name='Calibri',bold=True)
            shutter_width.alignment = copy(address.alignment)
            shutter_width.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['K'].width = ws.column_dimensions['K'].width*0.5


            #T
            shutter_thickness = ws.cell(row=10, column=12, value="T")
            shutter_thickness.font = Font(size=10,name='Calibri',bold=True)
            shutter_thickness.alignment = copy(address.alignment)
            shutter_thickness.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['L'].width = ws.column_dimensions['L'].width*0.5

            #Basic Rate/ Frame
            shutter_basic_rate = ws.cell(row=9, column=13, value="Basic Rate/\nShutter")
            shutter_basic_rate.font = Font(size=10,name='Calibri',bold=True)
            shutter_basic_rate.alignment = copy(address.alignment)
            shutter_basic_rate.fill=PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
            shutter_basic_rate.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=9, start_column=13, end_row=10, end_column=13)
            ws.column_dimensions[chr(ord('H'))].width = ws.column_dimensions[chr(ord('H'))].width*1.2


            gap_between = ws.cell(row=8, column=14)
            gap_between.fill = color5
            ws.merge_cells(start_row=8, start_column=14, end_row=10, end_column=14)
            ws.column_dimensions['N'].width = ws.column_dimensions['N'].width*0.1



            total_shutter_cost = ws.cell(row=8, column=15, value="Total Door Set Cost")
            total_shutter_cost.font = Font(size=10,name='Calibri',bold=True)
            total_shutter_cost.alignment = copy(address.alignment)
            total_shutter_cost.fill = color3
            total_shutter_cost.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=15, end_row=8, end_column=23 - disc_adj)

            frame_shutter = ws.cell(row=9, column=15, value="Basic Rate/\nFrame + Shutter")
            frame_shutter.font = Font(size=10,name='Calibri',bold=True)
            frame_shutter.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            frame_shutter.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            frame_shutter.fill =PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
            ws.merge_cells(start_row=9, start_column=15, end_row=10, end_column=15)
            ws.column_dimensions[chr(ord('M'))].width = ws.column_dimensions[chr(ord('M'))].width*1.2

            #Discount
            if disc_adj == 0:
                discount = ws.cell(row=9, column=16, value="Discount")
                discount.font = Font(size=10,name='Calibri',bold=True)
                discount.alignment = copy(address.alignment)
                discount.border = copy(serial_no.border)
                ws.merge_cells(start_row=9, start_column=16, end_row=9, end_column=17)
                #Percent
                percent = ws.cell(row=10, column=16, value="( % )")
                percent.font = Font(size=10,name='Calibri',bold=True)
                percent.alignment = copy(address.alignment)
                percent.border = copy(serial_no.border)
                #Amount
                amount = ws.cell(row=10, column=17, value="( Amt )")
                amount.font = Font(size=10,name='Calibri',bold=True)
                amount.alignment = copy(address.alignment)
                amount.border = copy(serial_no.border)
                #Discounted Price
                discounted_price = ws.cell(row=9, column=18, value="Discounted\nPrice")
                discounted_price.font = Font(size=10,name='Calibri',bold=True)
                discounted_price.alignment = copy(address.alignment)
                discounted_price.border = copy(serial_no.border)
                ws.merge_cells(start_row=9, start_column=18, end_row=10, end_column=18)
                ws.column_dimensions['R'].width = ws.column_dimensions['R'].width*0.9

            gst_percent = ws.cell(row=9, column=19 - disc_adj, value="GST")
            gst_percent.font = Font(size=10,name='Calibri',bold=True)
            gst_percent.alignment = copy(address.alignment)
            gst_percent.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=9, start_column=19 - disc_adj, end_row=9, end_column=20 - disc_adj)

            #Percent
            percent = ws.cell(row=10, column=19 - disc_adj, value="( % )")
            percent.font = copy(serial_no.font)
            percent.alignment = copy(address.alignment)
            percent.border = copy(serial_no.border)
            # ws.column_dimensions['F'].width = ws.column_dimensions['F'].width*0.5
            #Amount
            amount = ws.cell(row=10, column=20 - disc_adj, value="( Amt )")
            amount.font = copy(serial_no.font)
            amount.alignment = copy(address.alignment)
            amount.border = copy(serial_no.border)

            tot_cost = ws.cell(row=9, column=21 - disc_adj, value="Total cost of set inclusive of duties and taxes")
            tot_cost.font = Font(size=10,name='Calibri',bold=True)
            tot_cost.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            tot_cost.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            tot_cost.fill=PatternFill(start_color='C4D69B',end_color='C4D69B',fill_type='solid')
            ws.merge_cells(start_row=9, start_column=21 - disc_adj, end_row=10, end_column=21 - disc_adj)
            if disc_adj == 0:
                ws.column_dimensions[chr(ord('U'))].width = ws.column_dimensions[chr(ord('U'))].width*1.2
            else:
                ws.column_dimensions[chr(ord('R'))].width = ws.column_dimensions[chr(ord('R'))].width*1.2


            qty = ws.cell(row=9, column=22 - disc_adj, value="Qty")
            qty.font = Font(size=10,name='Calibri',bold=True)
            qty.alignment = copy(address.alignment)
            qty.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=9, start_column=22 - disc_adj, end_row=10, end_column=22 - disc_adj)

            tot_cost_with_tax = ws.cell(row=9, column=23 - disc_adj, value="Total cost of set inclusive of duties and taxes")
            tot_cost_with_tax.font = Font(size=10,name='Calibri',bold=True)
            tot_cost_with_tax.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            tot_cost_with_tax.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            tot_cost_with_tax.fill = PatternFill(start_color='C4D69B',end_color='C4D69B',fill_type='solid')
            ws.merge_cells(start_row=9, start_column=23 - disc_adj, end_row=10, end_column=23 - disc_adj)
            if disc_adj == 0:
                ws.column_dimensions[chr(ord('W'))].width = ws.column_dimensions[chr(ord('W'))].width*1.2
            else:
                ws.column_dimensions[chr(ord('T'))].width = ws.column_dimensions[chr(ord('T'))].width*1.2
            ws.row_dimensions[9].height = 25
            ws.row_dimensions[10].height = 25


            #Lines
            total_amount = 0
            current_row = 11
            sl_no = 1
            locale.setlocale(locale.LC_MONETARY, 'en_IN')
            def format(my_str,limit):
                wrapper = textwrap.TextWrapper(width=limit)
                string = wrapper.fill(text=my_str)
                return string
            
            for line in self.order_line:
                #Serial Number
                serial_no = ws.cell(row=current_row, column=1, value=sl_no)
                serial_no.font = Font(size=10,name='Calibri')
                serial_no.alignment = copy(address.alignment)
                serial_no.border = Border(bottom=thin,top=thin,right=thin,left=thin)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+3, end_column=1)

                #name
                name = ws.cell(row=current_row, column=2, value=line.product_id.name)
                name.font = Font(size=8,name='Calibri')
                name.alignment = copy(address.alignment)
                name.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row+3, end_column=2)

                #frame description
                frame_desc = ws.cell(row=current_row, column=3, value=line.frame_description)
                frame_desc.font = Font(size=8,name='Calibri')
                frame_desc.alignment = copy(address.alignment)
                frame_desc.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row+3, end_column=3)

                #L
                frame_outersize_length = ws.cell(row=current_row, column=4, value=line.frame_lengthx)
                frame_outersize_length.font = Font(size=8,name='Calibri')
                frame_outersize_length.alignment = copy(address.alignment)
                frame_outersize_length.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row+3, end_column=4)
       

                #W
                frame_outersize_width = ws.cell(row=current_row, column=5, value=line.frame_width)
                frame_outersize_width.font = Font(size=8,name='Calibri')
                frame_outersize_width.alignment = copy(address.alignment)
                frame_outersize_width.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row+3, end_column=5)


                #D
                frame_outersize_diameter = ws.cell(row=current_row, column=6, value=line.frame_diameter)
                frame_outersize_diameter.font = Font(size=8,name='Calibri')
                frame_outersize_diameter.alignment = copy(address.alignment)
                frame_outersize_diameter.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row+3, end_column=6)


                #T
                frame_outersize_thickness = ws.cell(row=current_row, column=7, value=line.frame_thickness)
                frame_outersize_thickness.font = Font(size=8,name='Calibri')
                frame_outersize_thickness.alignment = copy(address.alignment)
                frame_outersize_thickness.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row+3, end_column=7)

                #Basic Price
                frame_basic_rate = ws.cell(row=current_row, column=8, value=locale.currency(line.frame_price,grouping=True))
                frame_basic_rate.font = Font(size=8,name='Calibri',bold=True)
                frame_basic_rate.alignment = copy(address.alignment)
                frame_basic_rate.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=8, end_row=current_row+3, end_column=8)


                #shutter description
                frame_desc = ws.cell(row=current_row, column=3, value=line.frame_description)
                frame_desc.font = Font(size=8,name='Calibri')
                frame_desc.alignment = copy(address.alignment)
                frame_desc.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row+3, end_column=3)

                #under frame
                door_desc = ws.cell(row=current_row, column=9, value=line.shutter_description)
                door_desc.font = Font(size=8,name='Calibri')
                door_desc.alignment = copy(address.alignment)
                door_desc.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=9, end_row=current_row+3, end_column=9)

                #L
                shutter_length = ws.cell(row=current_row, column=10, value=line.shutter_lengthx)
                shutter_length.font = Font(size=8,name='Calibri')
                shutter_length.alignment = copy(address.alignment)
                shutter_length.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=10, end_row=current_row+3, end_column=10)
       

                #W
                shutter_width = ws.cell(row=current_row, column=11, value=line.shutter_width)
                shutter_width.font = Font(size=8,name='Calibri')
                shutter_width.alignment = copy(address.alignment)
                shutter_width.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=11, end_row=current_row+3, end_column=11)



                #T
                shutter_thickness = ws.cell(row=current_row, column=12, value=line.shutter_thickness)
                shutter_thickness.font = Font(size=8,name='Calibri')
                shutter_thickness.alignment = copy(address.alignment)
                shutter_thickness.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=12, end_row=current_row+3, end_column=12)

                #shutter Basic Price
                shutter_basic_rate = ws.cell(row=current_row, column=13, value=locale.currency(line.shutter_price,grouping=True))
                shutter_basic_rate.font = Font(size=8,name='Calibri',bold=True)
                shutter_basic_rate.alignment = copy(address.alignment)
                shutter_basic_rate.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=13, end_row=current_row+3, end_column=13)


                line_gap_between = ws.cell(row=current_row, column=14)
                line_gap_between.fill = color5
                ws.merge_cells(start_row=current_row, start_column=14, end_row=current_row+3, end_column=14)

                #Basic Rate/ Frame + Shutter
                frame_shutter = ws.cell(row=current_row, column=15, value=locale.currency(line.price_unit,grouping=True))
                frame_shutter.font = Font(size=8,name='Calibri',bold=True)
                frame_shutter.alignment = copy(address.alignment)
                frame_shutter.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=15, end_row=current_row+3, end_column=15)

                price = line.price_unit

                #Discount
                if disc_adj == 0:
                    discount = line.price_unit*line.discount/100
                    price = line.price_unit - discount
                    #Percent
                    percent = ws.cell(row=current_row, column=16, value=str(line.discount)+'%')
                    percent.font =  Font(size=8,name='Calibri')
                    percent.alignment = copy(address.alignment)
                    percent.border = copy(serial_no.border)
                    ws.merge_cells(start_row=current_row, start_column=16, end_row=current_row+3, end_column=16)
                    #Amount
                    amount = ws.cell(row=current_row, column=17, value=locale.currency(discount,grouping=True))
                    amount.alignment = copy(address.alignment)
                    amount.font = Font(size=8,name='Calibri',bold=True)
                    amount.border = copy(serial_no.border)
                    ws.merge_cells(start_row=current_row, start_column=17, end_row=current_row+3, end_column=17)
                    #Discounted Price
                    discounted_price = ws.cell(row=current_row, column=18, value=locale.currency(price,grouping=True))
                    discounted_price.alignment = copy(address.alignment)
                    discounted_price.font = Font(size=8,name='Calibri',bold=True)
                    discounted_price.border = copy(serial_no.border)
                    ws.merge_cells(start_row=current_row, start_column=18, end_row=current_row+3, end_column=18)

                tax_prcnt = sum([sum([child.amount for child in tax.children_tax_ids]) for tax in line.tax_id])
                tax = price * tax_prcnt / 100
                price += tax
                #Percent
                percent = ws.cell(row=current_row, column=19 - disc_adj, value=str(tax_prcnt)+'%')
                percent.font =  Font(size=8,name='Calibri')
                percent.alignment = copy(address.alignment)
                percent.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=19 - disc_adj , end_row=current_row+3, end_column=19 - disc_adj )
                #Amount
                amount = ws.cell(row=current_row, column=20 - disc_adj , value=locale.currency(tax,grouping=True))
                amount.font =  Font(size=8,name='Calibri')
                amount.alignment = copy(address.alignment)
                amount.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=20 - disc_adj , end_row=current_row+3, end_column=20 - disc_adj )

                #total
                tot_cost = ws.cell(row=current_row, column=21 - disc_adj, value=locale.currency(price,grouping=True))
                tot_cost.font = Font(size=8,name='Calibri',bold=True)
                tot_cost.alignment = copy(address.alignment)
                tot_cost.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=21 - disc_adj, end_row=current_row+3, end_column=21 - disc_adj)

                #quantity
                qty = ws.cell(row=current_row, column=22 - disc_adj, value=line.product_uom_qty)
                qty.font = Font(size=8,name='Calibri')
                qty.alignment = copy(address.alignment)
                qty.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=22 - disc_adj, end_row=current_row+3, end_column=22 - disc_adj)

                #toatal with quantity
                tot_cost_with_tax = ws.cell(row=current_row, column=23 - disc_adj, value=locale.currency(line.product_uom_qty*(price),grouping=True))
                tot_cost_with_tax.font = Font(size=8,name='Calibri',bold=True)
                tot_cost_with_tax.alignment = copy(address.alignment)
                tot_cost_with_tax.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=23 - disc_adj, end_row=current_row+3, end_column=23 - disc_adj)

                #update counters
                total_amount += price
                current_row += 4
                sl_no += 1


            #Grand Total
            grand_total = ws.cell(row=current_row, column=1, value="Grand Total")
            grand_total.font = Font(size=8,name='Calibri',bold=True)
            grand_total.alignment = Alignment(horizontal='right',vertical='center')
            grand_total.border = copy(serial_no.border)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=22 - disc_adj)
            #Grand Total Amount
            grand_total_amount = ws.cell(row=current_row, column=23 - disc_adj, value=locale.currency(self.amount_total,grouping=True))
            grand_total_amount.font = copy(grand_total.font)
            grand_total_amount.alignment = Alignment(horizontal='center',vertical='center')
            grand_total_amount.border = copy(serial_no.border)
            ws.merge_cells(start_row=current_row, start_column=23 - disc_adj, end_row=current_row, end_column=23 - disc_adj)
            #Footer

            footer_note = ws.cell(row=current_row+1, column=1, value='Note:'+self.note)
            footer_note.font = Font(size=10,name='Calibri',bold=True)
            footer_note.alignment = Alignment(horizontal='left',vertical='center')
            footer_note.border = copy(serial_no.border)
            ws.merge_cells(start_row=current_row+1, start_column=1, end_row=current_row+1, end_column=13) 

            footer_taxes = ws.cell(row=current_row+2, column=1, value='TAXES: Taxes as applicable at time of dispatch')
            footer_taxes.font = Font(size=10,name='Calibri',bold=True)
            footer_taxes.alignment = Alignment(horizontal='left',vertical='center')
            footer_taxes.border = Border(top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=current_row+2, start_column=1, end_row=current_row+2, end_column=13)

            footer_transport = ws.cell(row=current_row+3, column=1, value='TRANSPORT:  Extra at actual')
            footer_transport.font = Font(size=10,name='Calibri',bold=True)
            footer_transport.alignment = Alignment(horizontal='left',vertical='center')
            footer_transport.border = Border(right=thin,left=thin)
            ws.merge_cells(start_row=current_row+3, start_column=1, end_row=current_row+3, end_column=13) 

            footer_ac = ws.cell(row=current_row+4, column=1, value='ASSEMBLY CHARGES:NA')
            footer_ac.font = Font(size=10,name='Calibri',bold=True)
            footer_ac.alignment = Alignment(horizontal='left',vertical='center')
            footer_ac.border = Border(right=thin,left=thin)
            ws.merge_cells(start_row=current_row+4, start_column=1, end_row=current_row+4, end_column=13) 

            footer_payment = ws.cell(row=current_row+5, column=1, value='PAYMENT: '+self.payment_term_id.name)
            footer_payment.font = Font(size=10,name='Calibri',bold=True)
            footer_payment.alignment = Alignment(horizontal='left',vertical='center')
            footer_payment.border = Border(bottom=thin,right=thin,left=thin)
            ws.merge_cells(start_row=current_row+5, start_column=1, end_row=current_row+5, end_column=13) 

            footer_signature = ws.cell(row=current_row+1, column=14)
            footer_signature.font = Font(size=10,name='Calibri',bold=True)
            footer_signature.alignment = Alignment(horizontal='center',vertical='center')
            footer_signature.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=current_row+1, start_column=14, end_row=current_row+5, end_column=23 - disc_adj) 

            # img_footer = Image('/odoo14/prixgen_odoo14/mass_xls_reports/models/mas.jpg')
            decoded_img_footer = base64.b64decode(self.env.company.logo)
            img_footer = Image(io.BytesIO(decoded_img_footer))
            # h, w = img_footer.height, img_footer.width
            h, w = 94, 132
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            ws.add_image(img_footer,'P'+str(current_row+1))

            
            # footer_logo = ws['O'+str(current_row+1)]
            # print('************************',footer_logo)
            # footer_logo.border = Border(bottom=thin,left=thin,top=thin)
            # ws.merge_cells(start_row=current_row+1, start_column=14, end_row=current_row+5, end_column=18)

            




        #Material group type frame
        if self.material_group == 'frame':

            # img = Image('/odoo14/prixgen_odoo14/mass_xls_reports/models/mas.jpg')
            decoded_img = base64.b64decode(self.env.company.logo)
            img = Image(io.BytesIO(decoded_img))
            # h, w = img.height, img.width
            h, w = 94, 132
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            marker = AnchorMarker(col=0, colOff=coloffset, row=0, rowOff=rowoffset)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(img)
            logo = ws['A1']
            logo.border = Border(bottom=thin,left=thin,top=thin)
            ws.merge_cells(start_row=1, start_column=1, end_row=5, end_column=2)
            
            #Address
            address = ws.cell(row=1, column=3, value="#36, Mysore-Hunsur road,\nHootagalli, Belavadi Post,\nMysore -570018 India\nTel : +91 821 2404242\nFax : +91 821 2404343\ninfo@masfurn.com")
            address.font = Font(size=10,name='Calibri')
            address.alignment = Alignment(horizontal='center',vertical='center')
            address.border = Border(bottom=thin,top=thin)
            ws.merge_cells(start_row=1, start_column=3, end_row=5, end_column=6)
            
            #Title
            title = ws.cell(row=1, column=7, value="QUOTATION")
            title.font = Font(size=20,bold=True,name='Calibri')
            title.alignment = copy(address.alignment)
            title.border = copy(address.border)
            ws.merge_cells(start_row=1, start_column=7, end_row=5, end_column=15 - disc_adj)

            #HeadInfoRight
            #Quotation Number
            quotation_no = ws.cell(row=1,column=16 - disc_adj,value="QUOTATION NO: "+self.name)
            quotation_no.font = Font(size=10,name='Calibri')
            quotation_no.alignment = copy(address.alignment)
            quotation_no.border = Border(top=thin,right=thin)
            ws.merge_cells(start_row=1, start_column=16 - disc_adj, end_row=2, end_column=18 - disc_adj)
            #Date
            date = ws.cell(row=3,column=16 - disc_adj,value="Date: "+str(self.date_order))
            date.font = Font(size=10,name='Calibri')
            date.alignment = copy(address.alignment)
            date.border = Border(right=thin)
            ws.merge_cells(start_row=3, start_column=16 - disc_adj, end_row=3, end_column=18 - disc_adj)
            #Validity
            if self.validity_date:
                valid_days = str((self.validity_date - self.date_order.date()).days)
                validity = ws.cell(row=4, column=16 - disc_adj,value="Quote valid for "+valid_days+" days")
                validity.font = Font(size=10,name='Calibri',color=colors.RED)
                validity.alignment = copy(address.alignment)
                validity.border = Border(bottom=thin,right=thin)
            else:
                validity = ws.cell(row=4, column=16 - disc_adj)
                validity.border = Border(bottom=thin,right=thin)
            ws.merge_cells(start_row=4, start_column=16 - disc_adj, end_row=5, end_column=18 - disc_adj)


            #Analytic Account
            analytic = ws.cell(row=6, column=1, value=self.analytic_account_id.name)
            analytic.font = Font(size=15,name='Calibri',bold=True)
            analytic.alignment = copy(address.alignment)
            analytic.fill = color4
            analytic.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=6, start_column=1, end_row=7, end_column=18 - disc_adj)


            #serial number
            serial_no = ws.cell(row=8, column=1, value="SL.\nNo.")
            serial_no.font = Font(size=10,name='Calibri',bold=True)
            serial_no.alignment = copy(address.alignment)
            serial_no.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=1, end_row=9, end_column=1)
            ws.column_dimensions['A'].width = ws.column_dimensions['A'].width*0.5

            #Description
            description = ws.cell(row=8, column=2, value="Description")
            description.font = Font(size=10,name='Calibri',bold=True)
            description.alignment = copy(address.alignment)
            description.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=2)
            ws.column_dimensions['B'].width = ws.column_dimensions['B'].width*1.3

            #name
            name = ws.cell(row=9, column=2, value="Name")
            name.font = Font(size=10,name='Calibri',bold=True)
            name.alignment = copy(address.alignment)
            name.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=2)

            #frame Size 
            frame_size = ws.cell(row=8, column=3, value="Frame Size")
            frame_size.font = Font(size=10,name='Calibri',bold=True)
            frame_size.alignment = copy(address.alignment)
            frame_size.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=3, end_row=8, end_column=9)


            #L
            frame_length = ws.cell(row=9, column=3, value="L mm")
            frame_length.font = Font(size=10,name='Calibri',bold=True)
            frame_length.alignment = copy(address.alignment)
            frame_length.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['C'].width = ws.column_dimensions['C'].width*0.5
            # ws.merge_cells(start_row=9, start_column=4, end_row=9, end_column=7)

            frame_length_gap = ws.cell(row=9, column=4, value="")
            frame_length_gap.font = Font(size=10,name='Calibri',bold=True)
            frame_length_gap.alignment = copy(address.alignment)
            frame_length_gap.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['D'].width = ws.column_dimensions['D'].width*0.25

            #W
            frame_width = ws.cell(row=9, column=5, value="W mm")
            frame_width.font = Font(size=10,name='Calibri',bold=True)
            frame_width.alignment = copy(address.alignment)
            frame_width.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['E'].width = ws.column_dimensions['E'].width*0.5

            frame_width_gap = ws.cell(row=9, column=6, value="")
            frame_width_gap.font = Font(size=10,name='Calibri',bold=True)
            frame_width_gap.alignment = copy(address.alignment)
            frame_width_gap.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['F'].width = ws.column_dimensions['F'].width*0.25

            #D
            frame_diameter = ws.cell(row=9, column=7, value="D mm")
            frame_diameter.font = Font(size=10,name='Calibri',bold=True)
            frame_diameter.alignment = copy(address.alignment)
            frame_diameter.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['G'].width = ws.column_dimensions['G'].width*0.5

            frame_diameter_gap = ws.cell(row=9, column=8, value="")
            frame_diameter_gap.font = Font(size=10,name='Calibri',bold=True)
            frame_diameter_gap.alignment = copy(address.alignment)
            frame_diameter_gap.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['H'].width = ws.column_dimensions['H'].width*0.25

            #T
            frame_thickness = ws.cell(row=9, column=9, value="T mm")
            frame_thickness.font = Font(size=10,name='Calibri',bold=True)
            frame_thickness.alignment = copy(address.alignment)
            frame_thickness.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['I'].width = ws.column_dimensions['I'].width*0.5

            #Basic Rate
            basic_rate = ws.cell(row=8, column=10, value="Basic Rate")
            basic_rate.font = Font(size=10,name='Calibri',bold=True)
            basic_rate.alignment = copy(address.alignment)
            basic_rate.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=10, end_row=9, end_column=10)
            ws.column_dimensions['J'].width = ws.column_dimensions['J'].width * 1.2


            #Discount
            if disc_adj == 0:
                discount = ws.cell(row=8, column=11, value="Discount")
                discount.font = Font(size=10,name='Calibri',bold=True)
                discount.alignment = copy(address.alignment)
                discount.border = copy(serial_no.border)
                ws.merge_cells(start_row=8, start_column=11, end_row=8, end_column=12)
                #Percent
                percent = ws.cell(row=9, column=11, value="( % )")
                percent.font = Font(size=10,name='Calibri',bold=True)
                percent.alignment = copy(address.alignment)
                percent.border = copy(serial_no.border)
                #Amount
                amount = ws.cell(row=9, column=12, value="( Amt )")
                amount.font = Font(size=10,name='Calibri',bold=True)
                amount.alignment = copy(address.alignment)
                amount.border = copy(serial_no.border)
                #Discounted Price
                discounted_price = ws.cell(row=8, column=13, value="Discounted\nPrice")
                discounted_price.font = Font(size=10,name='Calibri',bold=True)
                discounted_price.alignment = copy(address.alignment)
                discounted_price.border = copy(serial_no.border)
                ws.merge_cells(start_row=8, start_column=13, end_row=9, end_column=13)
                # ws.column_dimensions['R'].width = ws.column_dimensions['R'].width*0.9

            qty = ws.cell(row=8, column=14 - disc_adj, value="Qty")
            qty.font = Font(size=10,name='Calibri',bold=True)
            qty.alignment = copy(address.alignment)
            qty.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=14 - disc_adj, end_row=9, end_column=14 - disc_adj)

            #Total Basic
            total_basic_rate = ws.cell(row=8, column=15 - disc_adj, value="Total Basic\n Rate")
            total_basic_rate.font = Font(size=10,name='Calibri',bold=True)
            total_basic_rate.alignment = copy(address.alignment)
            total_basic_rate.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=15 - disc_adj, end_row=9, end_column=15 - disc_adj)


            gst_percent = ws.cell(row=8, column=16 - disc_adj, value="GST")
            gst_percent.font = Font(size=10,name='Calibri',bold=True)
            gst_percent.alignment = copy(address.alignment)
            gst_percent.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=16 - disc_adj, end_row=8, end_column=17 - disc_adj)

            #Percent
            percent = ws.cell(row=9, column=16 - disc_adj, value="( % )")
            percent.font = copy(serial_no.font)
            percent.alignment = copy(address.alignment)
            percent.border = copy(serial_no.border)
            # ws.column_dimensions['F'].width = ws.column_dimensions['F'].width*0.5
            #Amount
            amount = ws.cell(row=9, column=17 - disc_adj, value="( Amt )")
            amount.font = copy(serial_no.font)
            amount.alignment = copy(address.alignment)
            amount.border = copy(serial_no.border)


            total_value_inc_tax = ws.cell(row=8, column=18 - disc_adj, value="Total Value \nInclusive of\n taxes")
            total_value_inc_tax.font = Font(size=10,name='Calibri',bold=True)
            total_value_inc_tax.alignment = copy(address.alignment)
            total_value_inc_tax.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=18 - disc_adj, end_row=9, end_column=18 - disc_adj)
            if disc_adj == 0:
                ws.column_dimensions['R'].width = ws.column_dimensions['R'].width*1.2
                ws.column_dimensions['O'].width = ws.column_dimensions['O'].width*1.1
                ws.column_dimensions['M'].width = ws.column_dimensions['M'].width*1.1
            else:
                ws.column_dimensions['O'].width = ws.column_dimensions['O'].width*1.2
                ws.column_dimensions['L'].width = ws.column_dimensions['L'].width*1.1

            ws.row_dimensions[8].height = 25
            ws.row_dimensions[9].height = 25


            total_amount = 0
            current_row = 10
            sl_no = 1
            locale.setlocale(locale.LC_MONETARY, 'en_IN')
            def format(my_str,limit):
                wrapper = textwrap.TextWrapper(width=limit)
                string = wrapper.fill(text=my_str)
                return string
            
            for line in self.order_line:
                #Serial Number
                serial_no = ws.cell(row=current_row, column=1, value=sl_no)
                serial_no.font = Font(size=10,name='Calibri')
                serial_no.alignment = copy(address.alignment)
                serial_no.border = Border(bottom=thin,top=thin,right=thin,left=thin)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+1, end_column=1)

                #name
                name = ws.cell(row=current_row, column=2, value=line.name)
                name.font = Font(size=8,name='Calibri')
                name.alignment = copy(address.alignment)
                name.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row+1, end_column=2)

                #L
                frame_length = ws.cell(row=current_row, column=3, value=line.frame_lengthx)
                frame_length.font = Font(size=8,name='Calibri')
                frame_length.alignment = copy(address.alignment)
                frame_length.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row+1, end_column=3)

                frame_length_gap = ws.cell(row=current_row, column=4, value='X')
                frame_length_gap.font = Font(size=8,name='Calibri')
                frame_length_gap.alignment = copy(address.alignment)
                frame_length_gap.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row+1, end_column=4)
       

                #W
                frame_width = ws.cell(row=current_row, column=5, value=line.frame_width)
                frame_width.font = Font(size=8,name='Calibri')
                frame_width.alignment = copy(address.alignment)
                frame_width.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row+1, end_column=5)

                frame_width_gap = ws.cell(row=current_row, column=6, value='X')
                frame_width_gap.font = Font(size=8,name='Calibri')
                frame_width_gap.alignment = copy(address.alignment)
                frame_width_gap.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row+1, end_column=6)


                #D
                frame_diameter = ws.cell(row=current_row, column=7, value=line.frame_diameter)
                frame_diameter.font = Font(size=8,name='Calibri')
                frame_diameter.alignment = copy(address.alignment)
                frame_diameter.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row+1, end_column=7)

                frame_diameter_gap = ws.cell(row=current_row, column=8, value='X')
                frame_diameter_gap.font = Font(size=8,name='Calibri')
                frame_diameter_gap.alignment = copy(address.alignment)
                frame_diameter_gap.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=8, end_row=current_row+1, end_column=8)


                #T
                frame_thickness = ws.cell(row=current_row, column=9, value=line.frame_thickness)
                frame_thickness.font = Font(size=8,name='Calibri')
                frame_thickness.alignment = copy(address.alignment)
                frame_thickness.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=9, end_row=current_row+1, end_column=9)

                #shutter Basic Price
                basic_rate = ws.cell(row=current_row, column=10, value=locale.currency(line.price_unit,grouping=True))
                basic_rate.font = Font(size=8,name='Calibri',bold=True)
                basic_rate.alignment = copy(address.alignment)
                basic_rate.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=10, end_row=current_row+1, end_column=10)

                price = line.price_unit


                 #Discount
                if disc_adj == 0:
                    discount = line.price_unit*line.discount/100
                    price = line.price_unit - discount
                    #Percent
                    percent = ws.cell(row=current_row, column=11, value=str(line.discount)+'%')
                    percent.font =  Font(size=8,name='Calibri')
                    percent.alignment = copy(address.alignment)
                    percent.border = copy(serial_no.border)
                    ws.merge_cells(start_row=current_row, start_column=11, end_row=current_row+1, end_column=11)
                    #Amount
                    amount = ws.cell(row=current_row, column=12, value=locale.currency(discount,grouping=True))
                    amount.alignment = copy(address.alignment)
                    amount.font = Font(size=8,name='Calibri',bold=True)
                    amount.border = copy(serial_no.border)
                    ws.merge_cells(start_row=current_row, start_column=12, end_row=current_row+1, end_column=12)
                    #Discounted Price
                    discounted_price = ws.cell(row=current_row, column=13, value=locale.currency(price,grouping=True))
                    discounted_price.alignment = copy(address.alignment)
                    discounted_price.font = Font(size=8,name='Calibri',bold=True)
                    discounted_price.border = copy(serial_no.border)
                    ws.merge_cells(start_row=current_row, start_column=13, end_row=current_row+1, end_column=13)

                #quantity
                qty = ws.cell(row=current_row, column=14 - disc_adj, value=line.product_uom_qty)
                qty.font = Font(size=8,name='Calibri')
                qty.alignment = copy(address.alignment)
                qty.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=14 - disc_adj, end_row=current_row+1, end_column=14 - disc_adj)

                #Total Basic Rate
                total_basic_rate = ws.cell(row=current_row, column=15 - disc_adj, value=locale.currency(line.price_subtotal,grouping=True))
                total_basic_rate.font = Font(size=8,name='Calibri',bold=True)
                total_basic_rate.alignment = copy(address.alignment)
                total_basic_rate.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=15 - disc_adj, end_row=current_row+1, end_column=15 - disc_adj)



                tax_prcnt = sum([sum([child.amount for child in tax.children_tax_ids]) for tax in line.tax_id])
                tax = price * tax_prcnt / 100
                price += tax

                #Percent
                percent = ws.cell(row=current_row, column=16 - disc_adj, value=str(tax_prcnt)+'%')
                percent.font =  Font(size=8,name='Calibri')
                percent.alignment = copy(address.alignment)
                percent.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=16 - disc_adj , end_row=current_row+1, end_column=16 - disc_adj )
                #Amount
                amount = ws.cell(row=current_row, column=17 - disc_adj , value=locale.currency(tax,grouping=True))
                amount.font =  Font(size=8,name='Calibri')
                amount.alignment = copy(address.alignment)
                amount.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=17 - disc_adj , end_row=current_row+1, end_column=17 - disc_adj )

                #total
                total_value_inc_tax = ws.cell(row=current_row, column=18 - disc_adj, value=locale.currency(line.product_uom_qty*(price),grouping=True))
                total_value_inc_tax.font = Font(size=8,name='Calibri',bold=True)
                total_value_inc_tax.alignment = copy(address.alignment)
                total_value_inc_tax.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=18 - disc_adj, end_row=current_row+1, end_column=18 - disc_adj)





                #update counters
                total_amount += price
                current_row += 2
                sl_no += 1


            #Grand Total
            grand_total = ws.cell(row=current_row, column=1, value="Grand Total")
            grand_total.font = Font(size=8,name='Calibri',bold=True)
            grand_total.alignment = Alignment(horizontal='right',vertical='center')
            grand_total.border = copy(serial_no.border)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=17 - disc_adj)
            #Grand Total Amount
            grand_total_amount = ws.cell(row=current_row, column=18 - disc_adj, value=locale.currency(self.amount_total,grouping=True))
            grand_total_amount.font = copy(grand_total.font)
            grand_total_amount.alignment = Alignment(horizontal='center',vertical='center')
            grand_total_amount.border = copy(serial_no.border)
            ws.merge_cells(start_row=current_row, start_column=18 - disc_adj, end_row=current_row, end_column=18 - disc_adj)
            #Footer

            footer_note = ws.cell(row=current_row+1, column=1, value='Note:'+self.note)
            footer_note.font = Font(size=10,name='Calibri',bold=True)
            footer_note.alignment = Alignment(horizontal='left',vertical='center')
            footer_note.border = copy(serial_no.border)
            ws.merge_cells(start_row=current_row+1, start_column=1, end_row=current_row+1, end_column=11) 

            footer_taxes = ws.cell(row=current_row+2, column=1, value='TAXES: Taxes as applicable at time of dispatch')
            footer_taxes.font = Font(size=10,name='Calibri',bold=True)
            footer_taxes.alignment = Alignment(horizontal='left',vertical='center')
            footer_taxes.border = Border(top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=current_row+2, start_column=1, end_row=current_row+2, end_column=11)

            footer_transport = ws.cell(row=current_row+3, column=1, value='TRANSPORT:  Extra at actual')
            footer_transport.font = Font(size=10,name='Calibri',bold=True)
            footer_transport.alignment = Alignment(horizontal='left',vertical='center')
            footer_transport.border = Border(right=thin,left=thin)
            ws.merge_cells(start_row=current_row+3, start_column=1, end_row=current_row+3, end_column=11) 

            footer_ac = ws.cell(row=current_row+4, column=1, value='FIXING: Rates do not include fixing charges')
            footer_ac.font = Font(size=10,name='Calibri',bold=True)
            footer_ac.alignment = Alignment(horizontal='left',vertical='center')
            footer_ac.border = Border(right=thin,left=thin)
            ws.merge_cells(start_row=current_row+4, start_column=1, end_row=current_row+4, end_column=11) 

            footer_payment = ws.cell(row=current_row+5, column=1, value='PAYMENT: '+self.payment_term_id.name)
            footer_payment.font = Font(size=10,name='Calibri',bold=True)
            footer_payment.alignment = Alignment(horizontal='left',vertical='center')
            footer_payment.border = Border(bottom=thin,right=thin,left=thin)
            ws.merge_cells(start_row=current_row+5, start_column=1, end_row=current_row+5, end_column=11) 


            # img_footer = Image('/odoo14/prixgen_odoo14/mass_xls_reports/models/mas.jpg')
            decoded_img_footer = base64.b64decode(self.env.company.logo)
            img_footer = Image(io.BytesIO(decoded_img_footer))
            # h, w = img_footer.height, img_footer.width
            h, w = 94, 132
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            ws.add_image(img_footer,'L'+str(current_row+1))

            foot_img = ws['L'+str(current_row+1)]
            foot_img.border = Border(bottom=thin,left=thin,top=thin)
            foot_img.alignment = Alignment(horizontal='center',vertical='center')
            ws.merge_cells(start_row=current_row+1, start_column=12, end_row=current_row+5, end_column=13)


            footer_signature = ws.cell(row=current_row+1, column=14)
            footer_signature.font = Font(size=10,name='Calibri',bold=True)
            footer_signature.alignment = Alignment(horizontal='center',vertical='center')
            footer_signature.border = Border(bottom=thin,top=thin,right=thin)
            ws.merge_cells(start_row=current_row+1, start_column=14, end_row=current_row+5, end_column=18 - disc_adj)




        #Material group type frame
        if self.material_group == 'furniture':
            picture_column = 0
            if self.with_pictures == False:
                picture_column = 1
            # img = Image('/odoo14/prixgen_odoo14/mass_xls_reports/models/mas.jpg')
            decoded_img = base64.b64decode(self.env.company.logo)
            img = Image(io.BytesIO(decoded_img))
            # h, w = img.height, img.width
            h, w = 94, 132
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            marker = AnchorMarker(col=0, colOff=coloffset, row=0, rowOff=rowoffset)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(img)
            logo = ws['A1']
            logo.border = Border(bottom=thin,left=thin,top=thin)
            ws.merge_cells(start_row=1, start_column=1, end_row=5, end_column=2)
            
            #Address
            address = ws.cell(row=1, column=3, value="#36, Mysore-Hunsur road,\nHootagalli, Belavadi Post,\nMysore -570018 India\nTel : +91 821 2404242\nFax : +91 821 2404343\ninfo@masfurn.com")
            address.font = Font(size=10,name='Calibri')
            address.alignment = Alignment(horizontal='center',vertical='center')
            address.border = Border(bottom=thin,top=thin)
            ws.merge_cells(start_row=1, start_column=3, end_row=5, end_column=6)
            
            #Title
            title = ws.cell(row=1, column=7, value="QUOTATION")
            title.font = Font(size=20,bold=True,name='Calibri')
            title.alignment = copy(address.alignment)
            title.border = copy(address.border)
            ws.merge_cells(start_row=1, start_column=7, end_row=5, end_column=15 - disc_adj -picture_column)

            #HeadInfoRight
            #Quotation Number
            quotation_no = ws.cell(row=1,column=16 - disc_adj -picture_column,value="QUOTATION NO: "+self.name)
            quotation_no.font = Font(size=10,name='Calibri')
            quotation_no.alignment = copy(address.alignment)
            quotation_no.border = Border(top=thin,right=thin)
            ws.merge_cells(start_row=1, start_column=16 - disc_adj -picture_column, end_row=2, end_column=18 - disc_adj -picture_column)
            #Date
            date = ws.cell(row=3,column=16 - disc_adj -picture_column,value="Date: "+str(self.date_order))
            date.font = Font(size=10,name='Calibri')
            date.alignment = copy(address.alignment)
            date.border = Border(right=thin)
            ws.merge_cells(start_row=3, start_column=16 - disc_adj -picture_column, end_row=3, end_column=18 - disc_adj -picture_column)
            #Validity
            if self.validity_date:
                valid_days = str((self.validity_date - self.date_order.date()).days)
                validity = ws.cell(row=4, column=16 - disc_adj -picture_column,value="Quote valid for "+valid_days+" days")
                validity.font = Font(size=10,name='Calibri',color=colors.RED)
                validity.alignment = copy(address.alignment)
                validity.border = Border(bottom=thin,right=thin)
            else:
                validity = ws.cell(row=4, column=16 - disc_adj -picture_column)
                validity.border = Border(bottom=thin,right=thin)
            ws.merge_cells(start_row=4, start_column=16 - disc_adj -picture_column, end_row=5, end_column=18 - disc_adj -picture_column)


            #Analytic Account
            analytic = ws.cell(row=6, column=1, value=self.analytic_account_id.name)
            analytic.font = Font(size=15,name='Calibri',bold=True)
            analytic.alignment = copy(address.alignment)
            analytic.fill = color4
            analytic.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=6, start_column=1, end_row=7, end_column=18 - disc_adj -picture_column)


            #serial number
            serial_no = ws.cell(row=8, column=1, value="SL.\nNo.")
            serial_no.font = Font(size=10,name='Calibri',bold=True)
            serial_no.alignment = copy(address.alignment)
            serial_no.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=1, end_row=9, end_column=1)
            ws.column_dimensions['A'].width = ws.column_dimensions['A'].width*0.5

            #code
            code = ws.cell(row=8, column=2, value="Code")
            code.font = Font(size=10,name='Calibri',bold=True)
            code.alignment = copy(address.alignment)
            code.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=2, end_row=9, end_column=2)
            # ws.column_dimensions['B'].width = ws.column_dimensions['B'].width*1.3

            #description
            description = ws.cell(row=8, column=3, value="Item description")
            description.font = Font(size=10,name='Calibri',bold=True)
            description.alignment = copy(address.alignment)
            description.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=3, end_row=9, end_column=5)

            if picture_column == 0:

                #picture
                picture = ws.cell(row=8, column=6, value="Picture")
                picture.font = Font(size=10,name='Calibri',bold=True)
                picture.alignment = copy(address.alignment)
                picture.border = Border(bottom=thin,top=thin,right=thin,left=thin)
                ws.merge_cells(start_row=8, start_column=6, end_row=9, end_column=6)
                ws.column_dimensions['F'].width = ws.column_dimensions['F'].width*1.5

            #Dimension/code
            funrniture_dimension = ws.cell(row=8, column=7-picture_column, value="Dimension/code")
            funrniture_dimension.font = Font(size=10,name='Calibri',bold=True)
            funrniture_dimension.alignment = copy(address.alignment)
            funrniture_dimension.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=7-picture_column, end_row=8, end_column=9-picture_column)

            ws.row_dimensions[9].height = 25
            ws.row_dimensions[8].height = 25


            #L
            furniture_length = ws.cell(row=9, column=7-picture_column, value="L")
            furniture_length.font = Font(size=10,name='Calibri',bold=True)
            furniture_length.alignment = copy(address.alignment)
            furniture_length.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['G'].width = ws.column_dimensions['G'].width*0.5
            # ws.merge_cells(start_row=9, start_column=4, end_row=9, end_column=7)

            #W
            furniture_width = ws.cell(row=9, column=8-picture_column, value="W")
            furniture_width.font = Font(size=10,name='Calibri',bold=True)
            furniture_width.alignment = copy(address.alignment)
            furniture_width.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['H'].width = ws.column_dimensions['H'].width*0.5

            #H
            furniture_height = ws.cell(row=9, column=9-picture_column, value="H")
            furniture_height.font = Font(size=10,name='Calibri',bold=True)
            furniture_height.alignment = copy(address.alignment)
            furniture_height.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.column_dimensions['I'].width = ws.column_dimensions['I'].width*0.5


            #Basic Rate/Unit
            basic_rate_unit = ws.cell(row=8, column=10-picture_column, value="Basic \nrate/unit")
            basic_rate_unit.font = Font(size=10,name='Calibri',bold=True)
            basic_rate_unit.alignment = copy(address.alignment)
            basic_rate_unit.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=10-picture_column, end_row=9, end_column=10-picture_column)

            #Discount
            if disc_adj == 0:
                discount = ws.cell(row=8, column=11-picture_column, value="Discount")
                discount.font = Font(size=10,name='Calibri',bold=True)
                discount.alignment = copy(address.alignment)
                discount.border = copy(serial_no.border)
                ws.merge_cells(start_row=8, start_column=11-picture_column, end_row=8, end_column=12-picture_column)
                #Percent
                percent = ws.cell(row=9, column=11-picture_column, value="( % )")
                percent.font = Font(size=10,name='Calibri',bold=True)
                percent.alignment = copy(address.alignment)
                percent.border = copy(serial_no.border)
                #Amount
                amount = ws.cell(row=9, column=12 - picture_column, value="( Amt )")
                amount.font = Font(size=10,name='Calibri',bold=True)
                amount.alignment = copy(address.alignment)
                amount.border = copy(serial_no.border)
                #Discounted Price
                discounted_price = ws.cell(row=8, column=13-picture_column, value="Discounted\nPrice")
                discounted_price.font = Font(size=10,name='Calibri',bold=True)
                discounted_price.alignment = copy(address.alignment)
                discounted_price.border = copy(serial_no.border)
                ws.merge_cells(start_row=8, start_column=13-picture_column, end_row=9, end_column=13-picture_column)
                # ws.column_dimensions['R'].width = ws.column_dimensions['R'].width*0.9

            #Total Basic
            total_basic_rate = ws.cell(row=8, column=14 - disc_adj -picture_column, value="Total Basic\n Rate")
            total_basic_rate.font = Font(size=10,name='Calibri',bold=True)
            total_basic_rate.alignment = copy(address.alignment)
            total_basic_rate.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=14 - disc_adj -picture_column, end_row=9, end_column=14 - disc_adj -picture_column)


            gst_percent = ws.cell(row=8, column=15 - disc_adj -picture_column, value="GST")
            gst_percent.font = Font(size=10,name='Calibri',bold=True)
            gst_percent.alignment = copy(address.alignment)
            gst_percent.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=15 - disc_adj -picture_column, end_row=8, end_column=16 - disc_adj -picture_column)

            #Percent
            percent = ws.cell(row=9, column=15 - disc_adj -picture_column, value="( % )")
            percent.font = copy(serial_no.font)
            percent.alignment = copy(address.alignment)
            percent.border = copy(serial_no.border)
            # ws.column_dimensions['F'].width = ws.column_dimensions['F'].width*0.5
            #Amount
            amount = ws.cell(row=9, column=16 - disc_adj -picture_column, value="( Amt )")
            amount.font = copy(serial_no.font)
            amount.alignment = copy(address.alignment)
            amount.border = copy(serial_no.border)


            total_value_inc_tax = ws.cell(row=8, column=17 - disc_adj -picture_column, value="Total Value \nInclusive of\n taxes")
            total_value_inc_tax.font = Font(size=10,name='Calibri',bold=True)
            total_value_inc_tax.alignment = copy(address.alignment)
            total_value_inc_tax.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=17 - disc_adj -picture_column, end_row=9, end_column=17 - disc_adj -picture_column)

            if disc_adj -picture_column == 0:
                ws.column_dimensions['N'].width = ws.column_dimensions['N'].width*1
               
            else:
                ws.column_dimensions['K'].width = ws.column_dimensions['K'].width*1

            qty = ws.cell(row=8, column=18 - disc_adj -picture_column, value="Qty")
            qty.font = Font(size=10,name='Calibri',bold=True)
            qty.alignment = copy(address.alignment)
            qty.border = Border(bottom=thin,top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=8, start_column=18 - disc_adj -picture_column, end_row=9, end_column=18 - disc_adj -picture_column)


            # ws.row_dimensions[8].height = 25
            # ws.row_dimensions[9].height = 25

            total_amount = 0
            current_row = 10
            sl_no = 1
            locale.setlocale(locale.LC_MONETARY, 'en_IN')
            def format(my_str,limit):
                wrapper = textwrap.TextWrapper(width=limit)
                string = wrapper.fill(text=my_str)
                return string
            
            for line in self.order_line:
                #Serial Number
                serial_no = ws.cell(row=current_row, column=1, value=sl_no)
                serial_no.font = Font(size=10,name='Calibri')
                serial_no.alignment = copy(address.alignment)
                serial_no.border = Border(bottom=thin,top=thin,right=thin,left=thin)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+1, end_column=1)

                code = ws.cell(row=current_row, column=2, value=line.name)
                code.font = Font(size=8,name='Calibri')
                code.alignment = copy(address.alignment)
                code.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row+1, end_column=2)

                description = ws.cell(row=current_row, column=3, value=line.name)
                description.font = Font(size=8,name='Calibri')
                description.alignment = copy(address.alignment)
                description.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row+1, end_column=5)
                # ws.row_dimensions[10].height = 25

                if picture_column == 0:
                    ws.row_dimensions[current_row].height = 50
                    ws.row_dimensions[current_row+1].height = 50

                    if line.product_id.image_1920:
                        # base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
                        # product_img = Image('%s/web/image/product.product/%s/image_1024' % (base_url,line.product_id.id))
                        productcellh = lambda x: c2e((x * 49.77)/99)
                        productcellw = lambda x: c2e((x * (18.65-1.71))/10)
                        productcoloffset = productcellw(0)
                        productrowoffset = productcellh(0)
                        decoded_img = base64.b64decode(line.product_id.image_1920)
                        product_img = Image(io.BytesIO(decoded_img))
                        h, w = 130,130
                        # h, w = product_img.height, product_img.width
                        size = XDRPositiveSize2D(p2e(w), p2e(h))
                        marker = AnchorMarker(col=5, colOff=0, row=current_row-1, rowOff=0)
                        product_img.anchor = OneCellAnchor(_from=marker, ext=size)
                        ws.add_image(product_img)
                        # ws.add_image(product_img,'F'+str(current_row+1))
                        prod_img = ws['F'+str(current_row)]
                        prod_img.border = Border(bottom=thin,left=thin,top=thin)
                        ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row+1, end_column=6)

                    # prod = ws['F'+str(current_row+1)]
                    # prod.border = Border(bottom=thin,left=thin,top=thin)
                    # prod.alignment = Alignment(horizontal='center',vertical='center')
                    # ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row+1, end_column=6)

                #L
                furniture_length = ws.cell(row=current_row, column=7-picture_column, value=line.lengthx)
                furniture_length.font = Font(size=8,name='Calibri')
                furniture_length.alignment = copy(address.alignment)
                furniture_length.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=7-picture_column, end_row=current_row+1, end_column=7-picture_column)
                
                #W
                furniture_width = ws.cell(row=current_row, column=8-picture_column, value=line.width)
                furniture_width.font = Font(size=8,name='Calibri')
                furniture_width.alignment = copy(address.alignment)
                furniture_width.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=8-picture_column, end_row=current_row+1, end_column=8-picture_column)
                
                #D
                furniture_height = ws.cell(row=current_row, column=9-picture_column, value=line.height)
                furniture_height.font = Font(size=8,name='Calibri')
                furniture_height.alignment = copy(address.alignment)
                furniture_height.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=9-picture_column, end_row=current_row+1, end_column=9-picture_column)

                basic_rate_unit = ws.cell(row=current_row, column=10-picture_column, value=locale.currency(line.price_unit,grouping=True))
                basic_rate_unit.font = Font(size=8,name='Calibri',bold=True)
                basic_rate_unit.alignment = copy(address.alignment)
                basic_rate_unit.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=10-picture_column, end_row=current_row+1, end_column=10-picture_column)

                price = line.price_unit

                #Discount
                if disc_adj == 0:
                    discount = line.price_unit*line.discount/100
                    price = line.price_unit - discount
                    #Percent
                    percent = ws.cell(row=current_row, column=11-picture_column, value=str(line.discount)+'%')
                    percent.font =  Font(size=8,name='Calibri')
                    percent.alignment = copy(address.alignment)
                    percent.border = copy(serial_no.border)
                    ws.merge_cells(start_row=current_row, start_column=11-picture_column, end_row=current_row+1, end_column=11-picture_column)
                    #Amount
                    amount = ws.cell(row=current_row, column=12 - picture_column, value=locale.currency(discount,grouping=True))
                    amount.alignment = copy(address.alignment)
                    amount.font = Font(size=8,name='Calibri',bold=True)
                    amount.border = copy(serial_no.border)
                    ws.merge_cells(start_row=current_row, start_column=12-picture_column, end_row=current_row+1, end_column=12-picture_column)
                    #Discounted Price
                    discounted_price = ws.cell(row=current_row, column=13-picture_column, value=locale.currency(price,grouping=True))
                    discounted_price.alignment = copy(address.alignment)
                    discounted_price.font = Font(size=8,name='Calibri',bold=True)
                    discounted_price.border = copy(serial_no.border)
                    ws.merge_cells(start_row=current_row, start_column=13-picture_column, end_row=current_row+1, end_column=13-picture_column)



                #Total Basic Rate
                total_basic_rate = ws.cell(row=current_row, column=14 - disc_adj -picture_column, value=locale.currency(line.price_subtotal,grouping=True))
                total_basic_rate.font = Font(size=8,name='Calibri',bold=True)
                total_basic_rate.alignment = copy(address.alignment)
                total_basic_rate.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=14 - disc_adj -picture_column, end_row=current_row+1, end_column=14 - disc_adj -picture_column)



                tax_prcnt = sum([sum([child.amount for child in tax.children_tax_ids]) for tax in line.tax_id])
                tax = price * tax_prcnt / 100
                price += tax

                #Percent
                percent = ws.cell(row=current_row, column=15 - disc_adj -picture_column, value=str(tax_prcnt)+'%')
                percent.font =  Font(size=8,name='Calibri')
                percent.alignment = copy(address.alignment)
                percent.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=15 - disc_adj -picture_column , end_row=current_row+1, end_column=15 - disc_adj -picture_column )
                #Amount
                amount = ws.cell(row=current_row, column=16 - disc_adj -picture_column , value=locale.currency(tax,grouping=True))
                amount.font =  Font(size=8,name='Calibri')
                amount.alignment = copy(address.alignment)
                amount.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=16 - disc_adj -picture_column , end_row=current_row+1, end_column=16 - disc_adj -picture_column )

                #total
                total_value_inc_tax = ws.cell(row=current_row, column=17 - disc_adj -picture_column, value=locale.currency(line.product_uom_qty*(price),grouping=True))
                total_value_inc_tax.font = Font(size=8,name='Calibri',bold=True)
                total_value_inc_tax.alignment = copy(address.alignment)
                total_value_inc_tax.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=17 - disc_adj -picture_column, end_row=current_row+1, end_column=17 - disc_adj -picture_column)

                #  #quantity
                qty = ws.cell(row=current_row, column=18 - disc_adj -picture_column, value=line.product_uom_qty)
                qty.font = Font(size=8,name='Calibri')
                qty.alignment = copy(address.alignment)
                qty.border = copy(serial_no.border)
                ws.merge_cells(start_row=current_row, start_column=18 - disc_adj -picture_column, end_row=current_row+1, end_column=18 - disc_adj -picture_column)

             #update counters
                total_amount += price
                current_row += 2
                sl_no += 1


            #Grand Total
            grand_total = ws.cell(row=current_row, column=1, value="Grand Total")
            grand_total.font = Font(size=8,name='Calibri',bold=True)
            grand_total.alignment = Alignment(horizontal='right',vertical='center')
            grand_total.border = copy(serial_no.border)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=17 - disc_adj -picture_column)
            #Grand Total Amount
            grand_total_amount = ws.cell(row=current_row, column=18 - disc_adj -picture_column, value=locale.currency(self.amount_total,grouping=True))
            grand_total_amount.font = copy(grand_total.font)
            grand_total_amount.alignment = Alignment(horizontal='center',vertical='center')
            grand_total_amount.border = copy(serial_no.border)
            ws.merge_cells(start_row=current_row, start_column=18 - disc_adj -picture_column, end_row=current_row, end_column=18 - disc_adj -picture_column)
            #Footer

            footer_note = ws.cell(row=current_row+1, column=1, value='Note:'+self.note)
            footer_note.font = Font(size=10,name='Calibri',bold=True)
            footer_note.alignment = Alignment(horizontal='left',vertical='center')
            footer_note.border = copy(serial_no.border)
            ws.merge_cells(start_row=current_row+1, start_column=1, end_row=current_row+1, end_column=11) 

            footer_taxes = ws.cell(row=current_row+2, column=1, value='TAXES: Taxes as applicable at time of dispatch')
            footer_taxes.font = Font(size=10,name='Calibri',bold=True)
            footer_taxes.alignment = Alignment(horizontal='left',vertical='center')
            footer_taxes.border = Border(top=thin,right=thin,left=thin)
            ws.merge_cells(start_row=current_row+2, start_column=1, end_row=current_row+2, end_column=11)

            footer_transport = ws.cell(row=current_row+3, column=1, value='TRANSPORT:  Extra at actual')
            footer_transport.font = Font(size=10,name='Calibri',bold=True)
            footer_transport.alignment = Alignment(horizontal='left',vertical='center')
            footer_transport.border = Border(right=thin,left=thin)
            ws.merge_cells(start_row=current_row+3, start_column=1, end_row=current_row+3, end_column=11) 

            footer_ac = ws.cell(row=current_row+4, column=1, value='FIXING: Rates do not include fixing charges')
            footer_ac.font = Font(size=10,name='Calibri',bold=True)
            footer_ac.alignment = Alignment(horizontal='left',vertical='center')
            footer_ac.border = Border(right=thin,left=thin)
            ws.merge_cells(start_row=current_row+4, start_column=1, end_row=current_row+4, end_column=11) 

            footer_payment = ws.cell(row=current_row+5, column=1, value='PAYMENT: '+self.payment_term_id.name)
            footer_payment.font = Font(size=10,name='Calibri',bold=True)
            footer_payment.alignment = Alignment(horizontal='left',vertical='center')
            footer_payment.border = Border(bottom=thin,right=thin,left=thin)
            ws.merge_cells(start_row=current_row+5, start_column=1, end_row=current_row+5, end_column=11) 


            # img_footer = Image('/odoo14/prixgen_odoo14/mass_xls_reports/models/mas.jpg')
            decoded_img_footer = base64.b64decode(self.env.company.logo)
            img_footer = Image(io.BytesIO(decoded_img_footer))
            # h, w = img_footer.height, img_footer.width
            h, w = 94, 132
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            ws.add_image(img_footer,'L'+str(current_row+1))

            foot_img = ws['L'+str(current_row+1)]
            foot_img.border = Border(bottom=thin,left=thin,top=thin)
            foot_img.alignment = Alignment(horizontal='center',vertical='center')
            ws.merge_cells(start_row=current_row+1, start_column=12, end_row=current_row+5, end_column=13)


            footer_signature = ws.cell(row=current_row+1, column=14)
            footer_signature.font = Font(size=10,name='Calibri',bold=True)
            footer_signature.alignment = Alignment(horizontal='center',vertical='center')
            footer_signature.border = Border(bottom=thin,top=thin,right=thin)
            ws.merge_cells(start_row=current_row+1, start_column=14, end_row=current_row+5, end_column=18 - disc_adj -picture_column) 
        






        fp = io.BytesIO()
        wb.save(fp)
        excel_file = base64.encodebytes(fp.getvalue())
        self.sale_quotation_xlsx_report = excel_file
        self.file_name = self.name+'_report.xlsx'
        fp.close()